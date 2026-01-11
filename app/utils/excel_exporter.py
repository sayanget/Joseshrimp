"""
Excel导出工具

提供各种报表的Excel导出功能
"""
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from flask import make_response
from io import BytesIO
from datetime import datetime

class ExcelExporter:
    """Excel导出器"""
    
    @staticmethod
    def create_workbook(title="Report"):
        """创建工作簿"""
        wb = Workbook()
        ws = wb.active
        ws.title = title
        return wb, ws
    
    @staticmethod
    def style_header(ws, row=1):
        """设置表头样式"""
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        for cell in ws[row]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
    
    @staticmethod
    def auto_adjust_column_width(ws):
        """自动调整列宽"""
        from openpyxl.cell.cell import MergedCell
        
        for column in ws.columns:
            max_length = 0
            column_letter = None
            
            for cell in column:
                # 跳过合并单元格
                if isinstance(cell, MergedCell):
                    continue
                    
                if column_letter is None:
                    column_letter = cell.column_letter
                    
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            if column_letter:
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
    
    @staticmethod
    def export_daily_sales(data, date_from, date_to):
        """导出每日销售报表"""
        wb, ws = ExcelExporter.create_workbook("Daily Sales")
        
        # 标题
        ws['A1'] = f'Daily Sales Report ({date_from} to {date_to})'
        ws.merge_cells('A1:F1')
        ws['A1'].font = Font(size=14, bold=True)
        ws['A1'].alignment = Alignment(horizontal="center")
        
        # 表头
        headers = ['Date', 'Orders', 'Total KG', 'Cash KG', 'Credit KG', 'Total Amount ($)']
        ws.append([])  # 空行
        ws.append(headers)
        ExcelExporter.style_header(ws, row=3)
        
        # 数据
        for row in data:
            ws.append([
                row['date'],
                row['order_count'],
                round(row['total_kg'], 2),
                round(row['cash_kg'], 2),
                round(row['credit_kg'], 2),
                round(row.get('total_amount', 0), 2)
            ])
        
        ExcelExporter.auto_adjust_column_width(ws)
        
        return ExcelExporter.create_response(wb, f'daily_sales_{date_from}_{date_to}.xlsx')
    
    @staticmethod
    def export_customer_sales(data, date_from=None, date_to=None):
        """导出客户销售排名"""
        wb, ws = ExcelExporter.create_workbook("Customer Sales")
        
        # 标题
        period = f' ({date_from} to {date_to})' if date_from and date_to else ''
        ws['A1'] = f'Customer Sales Ranking{period}'
        ws.merge_cells('A1:F1')
        ws['A1'].font = Font(size=14, bold=True)
        ws['A1'].alignment = Alignment(horizontal="center")
        
        # 表头
        headers = ['Rank', 'Customer', 'Orders', 'Total KG', 'Total Amount ($)', 'Last Sale']
        ws.append([])
        ws.append(headers)
        ExcelExporter.style_header(ws, row=3)
        
        # 数据
        for idx, row in enumerate(data, 1):
            ws.append([
                idx,
                row['customer_name'],
                row['order_count'],
                round(row['total_kg'], 2),
                round(row.get('total_amount', 0), 2),
                row.get('last_sale', '')
            ])
        
        ExcelExporter.auto_adjust_column_width(ws)
        
        return ExcelExporter.create_response(wb, f'customer_sales_{datetime.now().strftime("%Y%m%d")}.xlsx')
    
    @staticmethod
    def export_spec_sales(data, date_from=None, date_to=None):
        """导出规格使用统计"""
        wb, ws = ExcelExporter.create_workbook("Spec Sales")
        
        # 标题
        period = f' ({date_from} to {date_to})' if date_from and date_to else ''
        ws['A1'] = f'Specification Usage Ranking{period}'
        ws.merge_cells('A1:F1')
        ws['A1'].font = Font(size=14, bold=True)
        ws['A1'].alignment = Alignment(horizontal="center")
        
        # 表头
        headers = ['Rank', 'Specification', 'Usage Count', 'Total Boxes', 'Extra KG', 'Total KG']
        ws.append([])
        ws.append(headers)
        ExcelExporter.style_header(ws, row=3)
        
        # 数据
        for idx, row in enumerate(data, 1):
            ws.append([
                idx,
                row['spec_name'],
                row['usage_count'],
                row['total_boxes'],
                round(row.get('total_extra_kg', 0), 2),
                round(row['total_kg'], 2)
            ])
        
        ExcelExporter.auto_adjust_column_width(ws)
        
        return ExcelExporter.create_response(wb, f'spec_sales_{datetime.now().strftime("%Y%m%d")}.xlsx')
    
    @staticmethod
    def export_sales_by_representative(data, date_from=None, date_to=None):
        """导出销售员销售汇总"""
        wb, ws = ExcelExporter.create_workbook("Sales by Rep")
        
        # 标题
        period = f' ({date_from} to {date_to})' if date_from and date_to else ''
        ws['A1'] = f'Sales by Representative{period}'
        ws.merge_cells('A1:H1')
        ws['A1'].font = Font(size=14, bold=True)
        ws['A1'].alignment = Alignment(horizontal="center")
        
        # 表头
        headers = ['Rank', 'Representative', 'Orders', 'Total KG', 'Total Amount ($)', 
                  'Avg KG/Order', 'Cash KG', 'Credit KG']
        ws.append([])
        ws.append(headers)
        ExcelExporter.style_header(ws, row=3)
        
        # 数据
        for idx, row in enumerate(data, 1):
            ws.append([
                idx,
                row['representative'],
                row['order_count'],
                round(row['total_kg'], 2),
                round(row.get('total_amount', 0), 2),
                round(row.get('avg_kg_per_order', 0), 2),
                round(row.get('cash_kg', 0), 2),
                round(row.get('credit_kg', 0), 2)
            ])
        
        ExcelExporter.auto_adjust_column_width(ws)
        
        return ExcelExporter.create_response(wb, f'sales_by_representative_{datetime.now().strftime("%Y%m%d")}.xlsx')
    
    @staticmethod
    def export_representative_detail(data, representative, date_from=None, date_to=None):
        """导出指定销售员的销售记录详情"""
        wb, ws = ExcelExporter.create_workbook("Rep Detail")
        
        # 标题
        period = f' ({date_from} to {date_to})' if date_from and date_to else ''
        ws['A1'] = f'Sales Detail for {representative}{period}'
        ws.merge_cells('A1:G1')
        ws['A1'].font = Font(size=14, bold=True)
        ws['A1'].alignment = Alignment(horizontal="center")
        
        # 表头
        headers = ['Sale ID', 'Date', 'Customer', 'Payment Type', 'Total KG', 'Total Amount ($)', 'Status']
        ws.append([])
        ws.append(headers)
        ExcelExporter.style_header(ws, row=3)
        
        # 数据
        for row in data:
            customer_name = row.get('customer', {}).get('name', 'N/A') if row.get('customer') else 'N/A'
            ws.append([
                row['id'],
                row.get('sale_time', '').split('T')[0] if row.get('sale_time') else '',
                customer_name,
                row.get('payment_type', ''),
                round(row.get('total_kg', 0), 2),
                round(row.get('total_amount', 0), 2),
                row.get('status', '')
            ])
        
        ExcelExporter.auto_adjust_column_width(ws)
        
        return ExcelExporter.create_response(wb, f'sales_detail_{representative}_{datetime.now().strftime("%Y%m%d")}.xlsx')
    
    @staticmethod
    def create_response(wb, filename):
        """创建Flask响应"""
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        response = make_response(output.getvalue())
        response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        response.headers['Content-Disposition'] = f'attachment; filename={filename}'
        
        return response

def export_purchases_to_excel(purchases):
    """导出采购单列表"""
    wb, ws = ExcelExporter.create_workbook("Purchases")
    
    # 标题
    ws['A1'] = f'Purchase List ({datetime.now().strftime("%Y-%m-%d")})'
    ws.merge_cells('A1:F1')
    ws['A1'].font = Font(size=14, bold=True)
    ws['A1'].alignment = Alignment(horizontal="center")
    
    # 表头
    headers = ['Purchase ID', 'Purchase Time', 'Supplier', 'Total KG', 'Total Amount ($)', 'Payment Status']
    ws.append([])
    ws.append(headers)
    ExcelExporter.style_header(ws, row=3)
    
    # 数据
    for purchase in purchases:
        ws.append([
            purchase.id,
            purchase.purchase_time.strftime('%Y-%m-%d %H:%M'),
            purchase.supplier,
            round(float(purchase.total_kg), 3),
            round(float(purchase.total_amount), 2),
            purchase.payment_status
        ])
    
    ExcelExporter.auto_adjust_column_width(ws)
    
    return ExcelExporter.create_response(wb, f'purchases_{datetime.now().strftime("%Y%m%d")}.xlsx')
